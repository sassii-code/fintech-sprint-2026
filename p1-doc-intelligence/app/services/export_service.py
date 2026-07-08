import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _flatten(obj, prefix=""):
    out = {}
    for key, value in (obj or {}).items():
        path = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            out.update(_flatten(value, path))
        elif isinstance(value, list):
            out[path] = "; ".join(json.dumps(v) if isinstance(v, (dict, list)) else str(v) for v in value)
        else:
            out[path] = value
    return out


def build_workbook(records: list[dict]) -> io.BytesIO:
    """records: list of {"filename": str, "doc_type": str | None, "extracted_data": dict}"""
    flat_rows = []
    for rec in records:
        row = {"filename": rec.get("filename", ""), "doc_type": rec.get("doc_type") or ""}
        row.update(_flatten(rec.get("extracted_data") or {}))
        flat_rows.append(row)

    headers = []
    for row in flat_rows:
        for key in row:
            if key not in headers:
                headers.append(key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in flat_rows:
        ws.append([row.get(h, "") for h in headers])

    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = len(str(header))
        for row in flat_rows:
            max_len = max(max_len, len(str(row.get(header, ""))))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
